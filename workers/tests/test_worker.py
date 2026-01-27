"""
Integration tests for the Cloudflare Worker.

These tests run against a local instance of the worker using wrangler dev.
They verify the worker's HTTP responses, especially binary data handling.

Run with: pytest workers/tests/ -v
Requires: wrangler dev running on localhost:8788
"""

import pytest
import httpx
import zipfile
from io import BytesIO


# Worker URL - matches wrangler dev default port
WORKER_URL = "http://localhost:8788"


@pytest.fixture(scope="module")
def check_worker_running():
    """Check if the worker is running before running tests."""
    try:
        response = httpx.get(f"{WORKER_URL}/health", timeout=5.0)
        if response.status_code != 200:
            pytest.skip("Worker not running. Start with: cd workers && npx wrangler dev")
    except httpx.ConnectError:
        pytest.skip("Worker not running. Start with: cd workers && npx wrangler dev")


def create_mock_plan(athlete="Test Athlete", season="2025/2026"):
    """Create a valid 52-week plan for testing."""
    daily_patterns = {
        4: [3, 4, 1, 3, 4, 2, 0],
        3: [2, 3, 1, 3, 2, 1, 0],
        2: [2, 2, 1, 2, 2, 1, 0],
        1: [1, 2, 0, 1, 2, 0, 0],
        0: [0, 1, 0, 1, 0, 0, 0],
    }

    weeks = []
    for i in range(52):
        week_num = i + 1
        if week_num <= 3:
            phase, phase_type = "Transition", "taper"
        elif week_num <= 14:
            phase, phase_type = "General Prep I", "general-prep"
        elif week_num <= 18:
            phase, phase_type = "Special Prep I", "special-prep"
        elif week_num <= 27:
            phase, phase_type = "Competition I", "competition"
        elif week_num <= 28:
            phase, phase_type = "Transition", "taper"
        elif week_num <= 33:
            phase, phase_type = "General Prep II", "general-prep"
        elif week_num <= 37:
            phase, phase_type = "Special Prep II", "special-prep"
        elif week_num <= 45:
            phase, phase_type = "Competition II", "competition"
        else:
            phase, phase_type = "End of Season", "taper"

        load = 2 if phase_type == "taper" else 3

        weeks.append({
            "weekNum": week_num,
            "startDate": f"2025-{((i // 4) % 12) + 1:02d}-{(i % 4) * 7 + 1:02d}",
            "month": ["Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug"][(i // 4) % 12] + " 25",
            "phase": phase,
            "phaseType": phase_type,
            "block": f"Block {(i // 8) + 1}",
            "load": load,
            "competitions": [],
            "competitionImportance": None,
            "technical": "Event-specific drills",
            "physical": "Base building" if "Prep" in phase else "Maintenance",
            "dailyIntensity": daily_patterns.get(load, daily_patterns[2])
        })

    return {
        "athlete": athlete,
        "season": season,
        "eventGroup": "sprints",
        "periodization": "bi-phase",
        "seasonStart": "2025-08-25",
        "competitions": [],
        "weeks": weeks
    }


class TestWorkerHealth:
    """Basic health check tests."""

    def test_health_endpoint(self, check_worker_running):
        """Test that the worker health endpoint responds."""
        response = httpx.get(f"{WORKER_URL}/health")
        assert response.status_code == 200
        assert response.json()["status"] == "healthy"

    def test_root_endpoint(self, check_worker_running):
        """Test that the root endpoint returns service info."""
        response = httpx.get(f"{WORKER_URL}/")
        assert response.status_code == 200
        data = response.json()
        assert data["status"] == "ok"
        assert "athletics" in data["service"].lower()


class TestExcelDownload:
    """
    Tests for the Excel download endpoint.

    These tests specifically verify that binary data is correctly returned,
    preventing regressions like the Python bytes -> string conversion issue.
    """

    def test_excel_returns_valid_binary(self, check_worker_running):
        """
        CRITICAL: Test that Excel download returns actual binary data.

        This test catches the bug where Python bytes were converted to their
        string representation (b'PK...') instead of actual binary data.
        """
        plan = create_mock_plan()

        response = httpx.post(
            f"{WORKER_URL}/api/download-excel",
            json={"plan": plan},
            timeout=60.0
        )

        assert response.status_code == 200

        # Check content type
        assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        # CRITICAL: Verify it's actual binary, not string representation
        content = response.content

        # The bug would cause content to start with b"b'PK" (the string repr)
        # instead of the actual PK bytes
        assert not content.startswith(b"b'"), "Binary data was converted to string representation!"

        # Verify ZIP/XLSX magic bytes (PK = 0x504B)
        assert content[:2] == b"PK", f"Expected XLSX magic bytes 'PK', got {content[:10]!r}"
        assert content[:4] == b"PK\x03\x04", "Expected ZIP local file header signature"

    def test_excel_is_valid_zip(self, check_worker_running):
        """Test that the Excel file is a valid ZIP archive."""
        plan = create_mock_plan()

        response = httpx.post(
            f"{WORKER_URL}/api/download-excel",
            json={"plan": plan},
            timeout=60.0
        )

        assert response.status_code == 200

        # Try to open as ZIP
        try:
            with zipfile.ZipFile(BytesIO(response.content)) as zf:
                # XLSX files must contain these
                names = zf.namelist()
                assert "[Content_Types].xml" in names
                assert any("workbook.xml" in n for n in names)
                assert any("sheet1.xml" in n for n in names)
        except zipfile.BadZipFile:
            pytest.fail("Response is not a valid ZIP file - binary data may be corrupted")

    def test_excel_reasonable_size(self, check_worker_running):
        """Test that the Excel file has a reasonable size."""
        plan = create_mock_plan()

        response = httpx.post(
            f"{WORKER_URL}/api/download-excel",
            json={"plan": plan},
            timeout=60.0
        )

        assert response.status_code == 200

        # A 52-week plan should produce a file of reasonable size
        content_length = len(response.content)
        assert content_length > 1000, f"File too small ({content_length} bytes)"
        assert content_length < 1_000_000, f"File too large ({content_length} bytes)"

    def test_excel_filename_in_header(self, check_worker_running):
        """Test that Content-Disposition header has correct filename."""
        plan = create_mock_plan(athlete="Sprint Squad", season="2025/2026")

        response = httpx.post(
            f"{WORKER_URL}/api/download-excel",
            json={"plan": plan},
            timeout=60.0
        )

        assert response.status_code == 200

        content_disposition = response.headers.get("content-disposition", "")
        assert "attachment" in content_disposition
        assert ".xlsx" in content_disposition
        assert "Sprint" in content_disposition or "sprint" in content_disposition.lower()

    def test_excel_with_special_characters_in_name(self, check_worker_running):
        """Test Excel generation with special characters in athlete name."""
        plan = create_mock_plan(athlete="Test/Athlete: Squad!", season="2025/2026")

        response = httpx.post(
            f"{WORKER_URL}/api/download-excel",
            json={"plan": plan},
            timeout=60.0
        )

        assert response.status_code == 200
        assert response.content[:2] == b"PK"

        # Filename should be sanitized
        content_disposition = response.headers.get("content-disposition", "")
        assert "attachment" in content_disposition


    def test_excel_content_matches_plan(self, check_worker_running):
        """
        Test that the Excel file content matches the input plan data.

        This verifies the Excel is not just structurally valid, but contains
        the correct data from the plan.
        """
        from openpyxl import load_workbook

        plan = create_mock_plan(athlete="Test Runner", season="2025/2026")

        # Set specific values we can verify
        plan["weeks"][0]["phase"] = "Transition"
        plan["weeks"][0]["load"] = 2
        plan["weeks"][10]["phase"] = "General Prep I"
        plan["weeks"][10]["load"] = 3

        response = httpx.post(
            f"{WORKER_URL}/api/download-excel",
            json={"plan": plan},
            timeout=60.0
        )

        assert response.status_code == 200

        # Load the workbook
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active

        # Verify sheet name
        assert ws.title == "Annual Plan"

        # Verify title row (row 2, column B) contains athlete and season
        title_cell = ws.cell(row=2, column=2)
        assert "Test Runner" in str(title_cell.value)
        assert "2025/2026" in str(title_cell.value)

        # Verify week numbers are present (row 4)
        week_1_cell = ws.cell(row=4, column=2)  # Week 1
        assert week_1_cell.value == 1

        week_52_cell = ws.cell(row=4, column=53)  # Week 52
        assert week_52_cell.value == 52

        # Verify we have 52 weeks of data
        # Count non-empty cells in the week row (row 4, columns B through BA)
        week_count = sum(
            1 for col in range(2, 54)
            if ws.cell(row=4, column=col).value is not None
        )
        assert week_count == 52, f"Expected 52 weeks, found {week_count}"


class TestCORS:
    """Test CORS headers are properly set."""

    def test_cors_headers_on_options(self, check_worker_running):
        """Test that OPTIONS preflight returns CORS headers."""
        response = httpx.options(f"{WORKER_URL}/api/download-excel")

        assert response.status_code == 200
        assert "access-control-allow-origin" in response.headers
        assert "access-control-allow-methods" in response.headers

    def test_cors_headers_on_response(self, check_worker_running):
        """Test that responses include CORS headers."""
        response = httpx.get(f"{WORKER_URL}/health")

        assert response.status_code == 200
        assert "access-control-allow-origin" in response.headers
