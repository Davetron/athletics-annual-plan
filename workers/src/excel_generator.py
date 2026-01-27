"""
Excel generation service using openpyxl.
Creates formatted 52-week periodized training plans.
"""

from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Colors
COLORS = {
    "teal": PatternFill(start_color="13B5BD", end_color="13B5BD", fill_type="solid"),
    "orange": PatternFill(start_color="FF9900", end_color="FF9900", fill_type="solid"),
    "yellow": PatternFill(start_color="FFDD00", end_color="FFDD00", fill_type="solid"),
    "green": PatternFill(start_color="00CC66", end_color="00CC66", fill_type="solid"),
    "grey": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
    "light_blue": PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid"),
    "light_grey": PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"),
    # Training load colors
    "load_4": PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid"),
    "load_3": PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid"),
    "load_2": PatternFill(start_color="F1C40F", end_color="F1C40F", fill_type="solid"),
    "load_1": PatternFill(start_color="2ECC71", end_color="2ECC71", fill_type="solid"),
    "load_0": PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid"),
    # Competition importance colors
    "importance_1": PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid"),
    "importance_2": PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid"),
    "importance_3": PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid"),
}

PHASE_COLORS = {
    "general-prep": COLORS["orange"],
    "special-prep": COLORS["yellow"],
    "competition": COLORS["green"],
    "taper": COLORS["grey"],
}

LOAD_COLORS = {
    4: COLORS["load_4"],
    3: COLORS["load_3"],
    2: COLORS["load_2"],
    1: COLORS["load_1"],
    0: COLORS["load_0"],
}

IMPORTANCE_COLORS = {
    1: COLORS["importance_1"],
    2: COLORS["importance_2"],
    3: COLORS["importance_3"],
}

# Row definitions
ROWS = {
    "goals": 1,
    "title": 2,
    "month": 3,
    "week": 4,
    "week_commencing": 5,
    "competitions_label": 6,
    "importance": 7,
    "competition_detail": 8,
    "tests": 9,
    "monitoring": 10,
    "periods": 11,
    "phases": 12,
    "technical": 13,
    "tactical": 14,
    "physical": 15,
    "psychological": 16,
    "microcycles": 17,
    "block_name": 18,  # Block names row (merged cells)
    "block_intensity_4": 19,  # Intensity level 4 (red)
    "block_intensity_3": 20,  # Intensity level 3 (orange)
    "block_intensity_2": 21,  # Intensity level 2 (yellow)
    "block_intensity_1": 22,  # Intensity level 1 (green)
}

DATA_START_COL = 2  # Column B


def generate_excel_from_plan(plan: dict) -> Workbook:
    """
    Generate Excel workbook from plan data.

    Args:
        plan: Plan dictionary with athlete, season, weeks, etc.

    Returns:
        openpyxl Workbook object
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Annual Plan"

    weeks = plan.get("weeks", [])

    # Set column widths
    ws.column_dimensions["A"].width = 18
    for i in range(52):
        ws.column_dimensions[get_column_letter(DATA_START_COL + i)].width = 12

    # Add row labels
    _add_row_labels(ws)

    # Add title row
    _add_title_row(ws, plan)

    # Add month headers
    _add_month_headers(ws, weeks)

    # Add week numbers and dates
    _add_week_data(ws, weeks)

    # Add competitions
    _add_competitions(ws, weeks)

    # Add phases with colors
    _add_phases(ws, weeks)

    # Add training blocks (includes intensity visualization)
    _add_blocks(ws, weeks)

    # Add focus areas
    _add_focus_areas(ws, weeks)

    # Apply borders
    _apply_borders(ws, weeks)

    # Set row heights
    ws.row_dimensions[ROWS["competition_detail"]].height = 50

    return wb


def _add_row_labels(ws):
    """Add row labels in column A."""
    labels = {
        ROWS["goals"]: "Goals",
        ROWS["title"]: "Annual Plan",
        ROWS["month"]: "Month",
        ROWS["week"]: "Week",
        ROWS["week_commencing"]: "Week Commencing",
        ROWS["competitions_label"]: "Competitions",
        ROWS["importance"]: "Importance",
        ROWS["competition_detail"]: "Detail",
        ROWS["tests"]: "Tests",
        ROWS["monitoring"]: "Monitoring",
        ROWS["periods"]: "Periods",
        ROWS["phases"]: "Phases",
        ROWS["technical"]: "Technical",
        ROWS["tactical"]: "Tactical",
        ROWS["physical"]: "Physical",
        ROWS["psychological"]: "Psychological",
        ROWS["microcycles"]: "Microcycles",
        ROWS["block_name"]: "Block",
        ROWS["block_intensity_4"]: "",
        ROWS["block_intensity_3"]: "",
        ROWS["block_intensity_2"]: "",
        ROWS["block_intensity_1"]: "",
    }

    for row, label in labels.items():
        cell = ws.cell(row=row, column=1)
        cell.value = label
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(vertical="center")
        cell.fill = COLORS["light_grey"]

    # Merge and label the Weekly Load section (intensity rows)
    ws.merge_cells(
        start_row=ROWS["block_intensity_4"],
        start_column=1,
        end_row=ROWS["block_intensity_1"],
        end_column=1,
    )
    weekly_load_cell = ws.cell(row=ROWS["block_intensity_4"], column=1)
    weekly_load_cell.value = "Weekly Load"
    weekly_load_cell.font = Font(bold=True, size=10)
    weekly_load_cell.alignment = Alignment(vertical="center", horizontal="center")
    weekly_load_cell.fill = COLORS["light_grey"]


def _add_title_row(ws, plan: dict):
    """Add title row with athlete name and season."""
    cell = ws.cell(row=ROWS["title"], column=DATA_START_COL)
    cell.value = f"{plan.get('athlete', 'Unknown')} - {plan.get('season', 'Unknown')}"
    cell.font = Font(bold=True, size=14)
    cell.alignment = Alignment(horizontal="left", vertical="center")

    # Merge across several columns for the title
    ws.merge_cells(
        start_row=ROWS["title"],
        start_column=DATA_START_COL,
        end_row=ROWS["title"],
        end_column=DATA_START_COL + 10,
    )


def _add_month_headers(ws, weeks: list):
    """Add month headers with merged cells."""
    current_month = None
    month_start_col = DATA_START_COL

    for i, week in enumerate(weeks):
        col = DATA_START_COL + i
        week_month = week.get("month", "")

        if week_month != current_month:
            # If we had a previous month, merge it
            if current_month is not None and col > month_start_col:
                # Apply fill to all cells first
                for c in range(month_start_col, col):
                    month_cell = ws.cell(row=ROWS["month"], column=c)
                    month_cell.fill = COLORS["teal"]

                # Then merge
                if col - month_start_col > 1:
                    ws.merge_cells(
                        start_row=ROWS["month"],
                        start_column=month_start_col,
                        end_row=ROWS["month"],
                        end_column=col - 1,
                    )

                cell = ws.cell(row=ROWS["month"], column=month_start_col)
                cell.value = current_month
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(bold=True, color="FFFFFF")

            current_month = week_month
            month_start_col = col

    # Handle last month
    if current_month is not None:
        end_col = DATA_START_COL + len(weeks) - 1
        for c in range(month_start_col, end_col + 1):
            month_cell = ws.cell(row=ROWS["month"], column=c)
            month_cell.fill = COLORS["teal"]

        if end_col > month_start_col:
            ws.merge_cells(
                start_row=ROWS["month"],
                start_column=month_start_col,
                end_row=ROWS["month"],
                end_column=end_col,
            )

        cell = ws.cell(row=ROWS["month"], column=month_start_col)
        cell.value = current_month
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True, color="FFFFFF")


def _add_week_data(ws, weeks: list):
    """Add week numbers and commencing dates."""
    for i, week in enumerate(weeks):
        col = DATA_START_COL + i

        # Week number
        week_cell = ws.cell(row=ROWS["week"], column=col)
        week_cell.value = week.get("weekNum", i + 1)
        week_cell.alignment = Alignment(horizontal="center", vertical="center")
        week_cell.fill = COLORS["teal"]
        week_cell.font = Font(bold=True, color="FFFFFF")

        # Week commencing date
        date_cell = ws.cell(row=ROWS["week_commencing"], column=col)
        start_date = week.get("startDate", "")
        if start_date:
            try:
                dt = datetime.fromisoformat(start_date.replace("Z", "+00:00"))
                date_cell.value = dt.strftime("%d.%m.%y")
            except (ValueError, TypeError):
                date_cell.value = start_date
        date_cell.alignment = Alignment(horizontal="center", vertical="center")
        date_cell.font = Font(size=9)

        # Microcycles (repeat week number)
        micro_cell = ws.cell(row=ROWS["microcycles"], column=col)
        micro_cell.value = week.get("weekNum", i + 1)
        micro_cell.alignment = Alignment(horizontal="center", vertical="center")
        micro_cell.font = Font(size=9)


def _add_competitions(ws, weeks: list):
    """Add competition data."""
    for i, week in enumerate(weeks):
        col = DATA_START_COL + i
        competitions = week.get("competitions", [])

        if competitions:
            # Competition detail
            detail_cell = ws.cell(row=ROWS["competition_detail"], column=col)
            detail_cell.value = ", ".join(competitions)
            detail_cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            detail_cell.font = Font(size=9, bold=True)

            # Importance rating
            importance = week.get("competitionImportance")
            if importance:
                imp_cell = ws.cell(row=ROWS["importance"], column=col)
                imp_cell.value = importance
                imp_cell.alignment = Alignment(horizontal="center", vertical="center")
                imp_cell.fill = IMPORTANCE_COLORS.get(importance, COLORS["grey"])
                imp_cell.font = Font(bold=True, color="FFFFFF")


def _add_phases(ws, weeks: list):
    """Add phases with color coding."""
    current_phase = None
    phase_start_col = DATA_START_COL

    for i, week in enumerate(weeks):
        col = DATA_START_COL + i
        phase_name = week.get("phase", "")
        phase_type = week.get("phaseType", "")

        if phase_name != current_phase:
            # Merge previous phase cells
            if current_phase is not None and col > phase_start_col:
                prev_phase_type = weeks[i - 1].get("phaseType", "taper")
                color = PHASE_COLORS.get(prev_phase_type, COLORS["grey"])

                # Apply fill to all cells before merging
                for c in range(phase_start_col, col):
                    phase_cell = ws.cell(row=ROWS["phases"], column=c)
                    phase_cell.fill = color

                if col - phase_start_col > 1:
                    ws.merge_cells(
                        start_row=ROWS["phases"],
                        start_column=phase_start_col,
                        end_row=ROWS["phases"],
                        end_column=col - 1,
                    )

                cell = ws.cell(row=ROWS["phases"], column=phase_start_col)
                cell.value = current_phase
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(bold=True)

            current_phase = phase_name
            phase_start_col = col

    # Handle last phase
    if current_phase is not None:
        end_col = DATA_START_COL + len(weeks) - 1
        last_phase_type = weeks[-1].get("phaseType", "taper") if weeks else "taper"
        color = PHASE_COLORS.get(last_phase_type, COLORS["grey"])

        for c in range(phase_start_col, end_col + 1):
            phase_cell = ws.cell(row=ROWS["phases"], column=c)
            phase_cell.fill = color

        if end_col > phase_start_col:
            ws.merge_cells(
                start_row=ROWS["phases"],
                start_column=phase_start_col,
                end_row=ROWS["phases"],
                end_column=end_col,
            )

        cell = ws.cell(row=ROWS["phases"], column=phase_start_col)
        cell.value = current_phase
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True)


def _add_blocks(ws, weeks: list):
    """Add training blocks with intensity visualization.

    Creates a block section with:
    - Top row: Block names (merged cells, light blue background)
    - 4 rows below: Intensity levels (4, 3, 2, 1) showing weekly load
      in a stacked bar style visualization
    """
    current_block = None
    block_start_col = DATA_START_COL

    # First pass: Add block name headers (merged cells)
    for i, week in enumerate(weeks):
        col = DATA_START_COL + i
        block_name = week.get("block", "")

        if block_name != current_block:
            # Merge previous block cells
            if current_block is not None and col > block_start_col:
                for c in range(block_start_col, col):
                    block_cell = ws.cell(row=ROWS["block_name"], column=c)
                    block_cell.fill = COLORS["light_blue"]

                if col - block_start_col > 1:
                    ws.merge_cells(
                        start_row=ROWS["block_name"],
                        start_column=block_start_col,
                        end_row=ROWS["block_name"],
                        end_column=col - 1,
                    )

                cell = ws.cell(row=ROWS["block_name"], column=block_start_col)
                cell.value = current_block
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(bold=True)

            current_block = block_name
            block_start_col = col

    # Handle last block name
    if current_block is not None:
        end_col = DATA_START_COL + len(weeks) - 1
        for c in range(block_start_col, end_col + 1):
            block_cell = ws.cell(row=ROWS["block_name"], column=c)
            block_cell.fill = COLORS["light_blue"]

        if end_col > block_start_col:
            ws.merge_cells(
                start_row=ROWS["block_name"],
                start_column=block_start_col,
                end_row=ROWS["block_name"],
                end_column=end_col,
            )

        cell = ws.cell(row=ROWS["block_name"], column=block_start_col)
        cell.value = current_block
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True)

    # Second pass: Add intensity visualization rows (stacked bar style)
    # Map intensity level to its corresponding row
    intensity_rows = {
        4: ROWS["block_intensity_4"],
        3: ROWS["block_intensity_3"],
        2: ROWS["block_intensity_2"],
        1: ROWS["block_intensity_1"],
    }

    for i, week in enumerate(weeks):
        col = DATA_START_COL + i
        load = week.get("load", 2)  # Default to load 2 if not specified

        # Only display intensity if it's 1-4
        if load in intensity_rows:
            intensity_row = intensity_rows[load]
            cell = ws.cell(row=intensity_row, column=col)
            cell.value = load
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True)
            cell.fill = LOAD_COLORS.get(load, COLORS["load_2"])


def _add_focus_areas(ws, weeks: list):
    """Add technical and physical focus areas."""
    for i, week in enumerate(weeks):
        col = DATA_START_COL + i

        # Technical focus
        technical = week.get("technical", "")
        if technical:
            tech_cell = ws.cell(row=ROWS["technical"], column=col)
            tech_cell.value = technical
            tech_cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            tech_cell.font = Font(size=9)

        # Physical focus
        physical = week.get("physical", "")
        if physical:
            phys_cell = ws.cell(row=ROWS["physical"], column=col)
            phys_cell.value = physical
            phys_cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            phys_cell.font = Font(size=9)


def _apply_borders(ws, weeks: list):
    """Apply borders to the worksheet."""
    end_col = DATA_START_COL + len(weeks) - 1
    thin_border = Border(
        top=Side(style="thin", color="D0D0D0"),
        left=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
    )

    # Apply borders to all data cells
    for row in range(ROWS["month"], ROWS["block_intensity_1"] + 1):
        for col in range(1, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border

    # Add thicker borders between months
    thick_left = Side(style="medium", color="000000")
    current_month = None
    for i, week in enumerate(weeks):
        col = DATA_START_COL + i
        if week.get("month") != current_month and current_month is not None:
            # Add thick left border for month and week rows
            for row in [ROWS["month"], ROWS["week"]]:
                cell = ws.cell(row=row, column=col)
                cell.border = Border(
                    top=thin_border.top,
                    left=thick_left,
                    bottom=thin_border.bottom,
                    right=thin_border.right,
                )
        current_month = week.get("month")

    # Add thicker borders between blocks
    current_block = None
    for i, week in enumerate(weeks):
        col = DATA_START_COL + i
        if week.get("block") != current_block and current_block is not None:
            # Add thick left border for block and weekly load rows
            for row in range(ROWS["block_name"], ROWS["block_intensity_1"] + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = Border(
                    top=thin_border.top,
                    left=thick_left,
                    bottom=thin_border.bottom,
                    right=thin_border.right,
                )
        current_block = week.get("block")
